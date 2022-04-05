import requests
from lxml import etree
import json
import openpyxl
import time

outwp = openpyxl.Workbook()
outws = outwp.create_sheet(index=0)
outws.cell(row=1, column=1, value='index')
outws.cell(row=1, column=2, value='title')
outws.cell(row=1, column=3, value='url')
outws.cell(row=1, column=4, value='answer')
outws.cell(row=1, column=5, value='author')
count = 2
import urllib

headers = {
    "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36 SLBrowser/7.0.0.4071 SLBChan/30'
}


# 获取品论述
# def com(commonid ):
#
#     return comment
# #获取每一页数据
# def getlist(url):
#     global count
#
#     outws.cell(row=count, column=1, value=str(count))
#     outws.cell(row=count, column=2, value=str(title))
#     outws.cell(row=count, column=3, value=str(price))
#     outws.cell(row=count, column=4, value=str(comment_count))
#
#     count = count + 1


# getlist()
def getpage(i):
    url = f"https://bbs.hupu.com/bxj-{i}"
    res = requests.get(url, headers=headers)
    html = etree.HTML(res.text)
    title = html.xpath('//div[@class="post-title"]/a/text()')
    url = html.xpath('//div[@class="post-title"]/a/@href')
    author = html.xpath('//div[@class="post-auth"]/a/text()')
    answer = html.xpath('//div[@class="post-datum"]/text()')
    for title, url, answer, author in zip(title, url, answer, author):
        next_url = urllib.parse.urljoin('https://bbs.hupu.com/52891219.html', url)
        global count
        outws.cell(row=count, column=1, value=str(count))
        outws.cell(row=count, column=2, value=str(title))
        outws.cell(row=count, column=3, value=str(next_url))
        outws.cell(row=count, column=4, value=str(answer))
        outws.cell(row=count, column=5, value=str(author))
        count = count + 1


for i in range(20):
    getpage(i + 1)
    print(i + 1)
outwp.save("虎扑表.xlsx")
